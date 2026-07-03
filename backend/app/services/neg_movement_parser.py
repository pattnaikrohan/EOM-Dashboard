"""
Negative Movement Excel Parser — reads the 3-tab Negative Movement report from CargoWise.

Tab 1: "Negative movement>$250" — jobs where accrued cost ≠ final invoice by >$250
Tab 2: "Branch_Excess Profit > $5000" — jobs with unusually high profit
Tab 3: "Branch_Jobs with Losses" — closed jobs carrying a loss

All 3 tabs have a metadata header block (first ~6 rows) followed by a column header row and data.
"""
from __future__ import annotations
import re
from datetime import datetime, date
from openpyxl import load_workbook
from io import BytesIO

from app.services.staff_lookup import OPERATOR_NAMES, BRANCH_NAMES


# ── Consolidation / known-pattern keywords for auto-suppression ────────────────
_CONSOLIDATION_PATTERNS = re.compile(
    r'consolidat|consilid|consol\w*\s*parent|parent.*(?:consolidat|consilid|consol)|'
    r'parent\s*/?\s*consol|child\s*costs?\s*rolled|'
    r'charged\s+via\s+parent',
    re.IGNORECASE,
)

_KNOWN_PL_CATEGORIES = {
    "Accrued incorrectly",
    "Additional unbillable charges",
    "Consolidation parent job",
    "Exchange rate variance",
    "Dispute — awaiting credit",
    "Other",
}


def _auto_triage(comment: str, category: str, revenue: float) -> tuple[str, str]:
    """Determine resolution_status and category based on existing data.

    Returns (resolution_status, category).

    Rules (applied in priority order):
    1. If category already matches a known P&L reason  → "reviewed"
    2. If comment matches consolidation patterns AND revenue == 0 → "closed" + auto-set category
    3. If comment is non-empty (branch already responded) → "responded"
    4. Otherwise → "pending"
    """
    # Rule 1: Category already filled with a valid P&L reason
    if category and category in _KNOWN_PL_CATEGORIES:
        return "reviewed", category

    # Rule 2: Consolidation parent pattern detection
    if comment and _CONSOLIDATION_PATTERNS.search(comment) and revenue == 0:
        return "closed", "Consolidation parent job"

    # Rule 3: Branch already provided a comment
    if comment:
        return "responded", category

    # Rule 4: No comment, no category — needs operator input
    return "pending", category


def _parse_number(val) -> float:
    """Safely parse a number, handling comma-formatted strings."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("$", "")
    if not s or s == "-":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_date_str(val) -> str:
    """Convert a date value to DD/MM/YYYY string."""
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%d/%m/%Y")
    return str(val).strip()


def _find_header_row(rows: list, max_scan: int = 30) -> tuple[int, list[str]]:
    """Scan the first N rows to find the column header row.
    Returns (row_index, list_of_header_strings).
    Heuristic: a header row contains 'Job Number' or 'Job Local Ref'.
    """
    for i, row in enumerate(rows[:max_scan]):
        row_strs = [str(c or "").strip() for c in row]
        joined = " ".join(row_strs).lower()
        if "job number" in joined or "job local ref" in joined:
            return i, row_strs
    return -1, []


def _build_col_map(headers: list[str]) -> dict[str, int]:
    """Build a column-name → index map using flexible substring matching."""
    col_map = {}
    for idx, h in enumerate(headers):
        hl = h.lower()
        if "job number" in hl and "job_number" not in col_map:
            col_map["job_number"] = idx
        elif "local ref" in hl:
            col_map["job_local_ref"] = idx
        elif hl in ("br", "brn.", "brn", "branch", "job brn."):
            col_map["branch"] = idx
        elif "dept" in hl:
            col_map["department"] = idx
        elif "stat" in hl and "status" not in hl and "invoice" not in hl:
            col_map["status"] = idx
        elif "status" in hl or "invoice status" in hl:
            col_map["status"] = idx
        elif "trans" in hl and "transaction" not in hl:
            col_map["transport"] = idx
        elif "cont" in hl and "client" not in hl and "controlling" not in hl:
            col_map["container"] = idx
        elif "sales" in hl:
            col_map["sales_rep"] = idx
        elif "local client" in hl and "name" not in hl:
            col_map["local_client"] = idx
        elif "client name" in hl or "local client name" in hl:
            col_map["local_client_name"] = idx
        elif "origin" in hl:
            col_map["origin"] = idx
        elif "dest" in hl:
            col_map["destination"] = idx
        elif hl == "etd" or "etd" in hl:
            col_map["etd"] = idx
        elif hl == "eta" or "eta" in hl:
            col_map["eta"] = idx
        elif "profit" in hl:
            col_map["job_profit"] = idx
        elif "revenue" in hl:
            col_map["revenue"] = idx
        elif hl == "wip" or "wip" in hl:
            col_map["wip"] = idx
        elif "cost" in hl:
            col_map["cost"] = idx
        elif "accrual" in hl:
            col_map["accrual"] = idx
        elif "comment" in hl:
            col_map["comment"] = idx
        elif "category" in hl:
            col_map["category"] = idx
        elif "notes" in hl or "notes_ho" in hl:
            col_map["notes_ho"] = idx
        elif "opened" in hl or "open" in hl:
            col_map["opened"] = idx
    return col_map


def _parse_tab(ws, section_name: str) -> list[dict]:
    """Parse a single worksheet tab into a list of job dictionaries."""
    rows = list(ws.iter_rows(values_only=True))
    header_idx, headers = _find_header_row(rows)
    
    if header_idx < 0:
        return []
    
    col_map = _build_col_map(headers)
    jobs = []
    now_iso = datetime.now().isoformat()
    
    for row in rows[header_idx + 1:]:
        if not row:
            continue
        
        # Get job number
        jn_idx = col_map.get("job_number", 0)
        job_num = str(row[jn_idx] or "").strip() if jn_idx < len(row) else ""
        
        # Skip empty rows, summary rows, and rows that don't look like job numbers
        if not job_num or not re.match(r'^[A-Z]\d{5,}', job_num):
            continue
        
        def _g(key: str):
            idx = col_map.get(key)
            if idx is not None and idx < len(row):
                return row[idx]
            return None
        
        branch_code = str(_g("branch") or "").strip()
        sales_code = str(_g("sales_rep") or "").strip()
        
        job = {
            "job_number":     job_num,
            "job_local_ref":  str(_g("job_local_ref") or "").strip(),
            "branch":         BRANCH_NAMES.get(branch_code, branch_code),
            "department":     str(_g("department") or "").strip(),
            "status":         str(_g("status") or "").strip(),
            "transport":      str(_g("transport") or "").strip(),
            "container":      str(_g("container") or "").strip(),
            "sales_rep":      OPERATOR_NAMES.get(sales_code, sales_code),
            "local_client":   str(_g("local_client") or _g("local_client_name") or "").strip(),
            "origin":         str(_g("origin") or "").strip(),
            "destination":    str(_g("destination") or "").strip(),
            "etd":            _parse_date_str(_g("etd") or _g("opened")),
            "eta":            _parse_date_str(_g("eta")),
            "job_profit":     _parse_number(_g("job_profit")),
            "revenue":        _parse_number(_g("revenue")),
            "wip":            _parse_number(_g("wip")),
            "cost":           _parse_number(_g("cost")),
            "accrual":        _parse_number(_g("accrual")),
            "section":        section_name,
            # Commentary fields — preserve existing comments from Excel
            "comment":        str(_g("comment") or "").strip(),
            "category":       str(_g("category") or "").strip(),
            "notes_ho":       str(_g("notes_ho") or "").strip(),
            # Workflow fields (placeholder — overwritten by auto-triage below)
            "resolution_status": "pending",
            "assigned_to":    OPERATOR_NAMES.get(sales_code, sales_code),
            "updated_at":     now_iso,
            "created_at":     now_iso,
        }
        
        # Run auto-triage to set resolution_status and possibly update category
        triage_status, triage_category = _auto_triage(
            job["comment"], job["category"], job["revenue"]
        )
        job["resolution_status"] = triage_status
        if triage_category:
            job["category"] = triage_category
        
        jobs.append(job)
    
    return jobs


# Section identifiers and their matching patterns
SECTION_PATTERNS = [
    ("negative_movement", ["negative movement", "negative"]),
    ("excess_profit",     ["excess profit", "excess"]),
    ("jobs_with_losses",  ["jobs with losses", "losses", "loss"]),
]


def _detect_section(sheet_name: str) -> str:
    """Determine which section a sheet belongs to based on its name."""
    name_lower = sheet_name.lower()
    for section_id, patterns in SECTION_PATTERNS:
        for pattern in patterns:
            if pattern in name_lower:
                return section_id
    return "negative_movement"  # default fallback


def parse_neg_movement_excel(file_bytes: bytes) -> dict:
    """
    Parse a Negative Movement Excel report (3 tabs).
    Returns { branch, period, sections: { section_name: [jobs] } }
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    
    all_jobs: dict[str, list[dict]] = {
        "negative_movement": [],
        "excess_profit": [],
        "jobs_with_losses": [],
    }
    
    branch = ""
    period = ""
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        section = _detect_section(sheet_name)
        
        # Try to extract branch and period from the first few rows
        rows = list(ws.iter_rows(values_only=True, max_row=10))
        for row in rows[:6]:
            for cell in row:
                cell_str = str(cell or "").strip()
                # Look for period info like "To: 2026-01-31"
                if "to:" in cell_str.lower():
                    try:
                        date_part = cell_str.split(":")[-1].strip()
                        d = datetime.strptime(date_part, "%Y-%m-%d")
                        period = d.strftime("%B %Y")
                    except (ValueError, IndexError):
                        pass
                # Look for date cells that indicate period
                if isinstance(cell, (datetime, date)) and not period:
                    period = cell.strftime("%B %Y")
        
        jobs = _parse_tab(ws, section)
        all_jobs[section].extend(jobs)
        
        # Extract branch from first job
        if not branch and jobs:
            branch = jobs[0]["branch"]
    
    wb.close()
    
    return {
        "branch": branch or "ALL",
        "period": period or datetime.now().strftime("%B %Y"),
        "sections": all_jobs,
    }
