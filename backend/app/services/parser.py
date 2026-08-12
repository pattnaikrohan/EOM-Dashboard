"""
Excel Parser — reads CargoWise export and WIP Review files.

Handles two formats:
  1. CargoWise raw export (single 'Shipment Profile' sheet with report header)
  2. WIP Review output (per-operator sheets with section banners)
"""
from __future__ import annotations
import re
from datetime import datetime, date
from openpyxl import load_workbook
from io import BytesIO

from app.services.rules import get_flags, priority_flag, get_ops_section, is_export_dept
from app.services.staff_lookup import OPERATOR_NAMES, BRANCH_NAMES, normalize_branch_name



def _parse_number(val) -> float:
    """Safely parse a number, handling comma-formatted strings."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("$", "")
    if not s or s == "-":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_date_str(val) -> str:
    """Convert a date value to DD/MM/YYYY string."""
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%d/%m/%Y")
    return str(val).strip()


def _compute_age(val) -> int:
    """Compute days since open date."""
    if val is None:
        return 0
    today = datetime.now()
    if isinstance(val, datetime):
        return max(0, (today - val).days)
    if isinstance(val, date):
        return max(0, (today.date() - val).days)
    # Try parsing common formats
    for fmt in ("%d/%m/%Y", "%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d"):
        try:
            d = datetime.strptime(str(val).strip(), fmt)
            return max(0, (today - d).days)
        except ValueError:
            continue
    return 0


def parse_wip_review(file_bytes: bytes) -> dict:
    """
    Parse a WIP Review Excel file (per-operator sheets).
    Returns { branch, period, operators: [str], jobs: [dict] }
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    
    skip_sheets = {"Legend", "Ops Manager Review"}
    all_jobs = []
    operators = []
    branch = ""
    period = ""
    
    # ── Parse operator sheets ──────────────────────────────────────────────
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            operators.append(sheet_name)
            continue
        
        # Row 0 = title like "WIP Review - DL2 - May 2026"
        title_row = rows[0]
        title = str(title_row[0] or "") if title_row else ""
        
        # Extract period from title
        match = re.search(r"-\s*(\w+\s+\d{4})\s*$", title)
        if match and not period:
            period = match.group(1)
        
        # Row 1 = headers
        headers = [str(h or "").strip() for h in rows[1]]
        
        # Build column index map
        col_map = {}
        for idx, h in enumerate(headers):
            hl = h.lower()
            if "job number" in hl:
                col_map["job_number"] = idx
            elif "job status" in hl:
                col_map["job_status"] = idx
            elif "branch" in hl:
                col_map["branch"] = idx
            elif "department" in hl:
                col_map["department"] = idx
            elif "open date" in hl:
                col_map["open_date"] = idx
            elif "operator" in hl:
                col_map["operator"] = idx
            elif "sales rep" in hl:
                col_map["sales_rep"] = idx
            elif "local charges" in hl:
                col_map["local_charges"] = idx
            elif "overseas agent" in hl:
                col_map["overseas_agent"] = idx
            elif "revenue" in hl:
                col_map["revenue"] = idx
            elif hl == "wip":
                col_map["wip"] = idx
            elif "cost" in hl:
                col_map["cost"] = idx
            elif "accrual" in hl:
                col_map["accrual"] = idx
            elif "profit" in hl or "p/l" in hl or "profit/loss" in hl:
                col_map["profit_loss"] = idx
            elif "margin" in hl:
                col_map["margin_pct"] = idx
            elif "age" in hl:
                col_map["job_age_days"] = idx
            elif "flag" in hl:
                col_map["flags_str"] = idx
            elif "etd" in hl or "origin etd" in hl or "departure" in hl:
                col_map["etd"] = idx
            elif "eta" in hl or "destination eta" in hl or "arrival" in hl:
                col_map["eta"] = idx
        
        operators.append(sheet_name)
        
        # Parse data rows (skip section banners)
        for row in rows[2:]:
            # Section banners have empty first cell and text in second
            job_num_idx = col_map.get("job_number", 0)
            job_num = str(row[job_num_idx] or "").strip() if row[job_num_idx] else ""
            
            # Skip empty rows and section banner rows
            if not job_num or not re.match(r'^[A-Z]\d{5,}', job_num):
                continue
            
            def _get(key: str):
                idx = col_map.get(key)
                if idx is not None and idx < len(row):
                    return row[idx]
                return None
            
            open_date_raw = _get("open_date")
            dept = str(_get("department") or "").strip()
            age_raw = _get("job_age_days")
            age = int(_parse_number(age_raw)) if age_raw is not None else _compute_age(open_date_raw)
            
            branch_code = str(_get("branch") or "").strip()
            op_code = str(_get("operator") or sheet_name).strip()
            op = OPERATOR_NAMES.get(op_code, op_code)
            job_branch = normalize_branch_name(branch_code)
            
            job = {
                "job_number":     job_num,
                "job_status":     str(_get("job_status") or "").strip(),
                "branch":         job_branch,
                "department":     dept,
                "open_date":      _parse_date_str(open_date_raw),
                "operator":       op,

                "sales_rep":      str(_get("sales_rep") or "").strip(),
                "local_charges":  str(_get("local_charges") or "").strip(),
                "overseas_agent": str(_get("overseas_agent") or "").strip(),
                "revenue":        _parse_number(_get("revenue")),
                "wip":            _parse_number(_get("wip")),
                "cost":           _parse_number(_get("cost")),
                "accrual":        _parse_number(_get("accrual")),
                "profit_loss":    _parse_number(_get("profit_loss")),
                "margin_pct":     _parse_number(_get("margin_pct")),
                "job_age_days":   age,
                "is_export":      is_export_dept(dept),
                "etd":            _parse_date_str(_get("etd") or open_date_raw),
                "eta":            _parse_date_str(_get("eta") or open_date_raw),
            }
            
            # Get branch
            if job["branch"] and not branch:
                branch = job["branch"]
            
            # Compute flags from data (or use provided flags)
            flags_str = str(_get("flags_str") or "").strip()
            if flags_str:
                job["flags"] = [f.strip() for f in flags_str.split(",") if f.strip()]
            else:
                job["flags"] = get_flags(job, period)
            
            job["primary_flag"] = priority_flag(job["flags"])
            job["ops_section"] = get_ops_section(job)
            
            all_jobs.append(job)
    
    wb.close()
    
    return {
        "branch": BRANCH_NAMES.get(branch, branch) if branch else "ALL",
        "period": period or datetime.now().strftime("%B %Y"),
        "operators": [OPERATOR_NAMES.get(op, op) for op in operators],
        "jobs": all_jobs,
    }


def parse_cargowise_export(file_bytes: bytes, **kwargs) -> dict:
    """
    Parse a raw CargoWise Shipment Profile export.
    Returns same structure as parse_wip_review.
    source_type: one of 'exports', 'imports_b', 'imports_s', 'cross_trade', or '' (unknown).
    """
    source_type = kwargs.get("source_type", "")
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    
    # Find the Shipment Profile sheet
    target_sheet = None
    for name in wb.sheetnames:
        if "shipment" in name.lower() or "profile" in name.lower():
            target_sheet = name
            break
    if not target_sheet:
        target_sheet = wb.sheetnames[0]
    
    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    
    # Scan for header row (contains "Shipment ID" or "Job Number")
    header_idx = None
    branch = ""
    for i, row in enumerate(rows[:30]):
        row_strs = [str(c or "").strip() for c in row]
        # Check for branch
        for cell in row_strs:
            if "job branch:" in cell.lower():
                parts = cell.split(":")
                if len(parts) > 1:
                    branch = parts[1].strip().split(",")[0].strip()
        # Check for header
        if any(s.lower() in ("shipment id", "job number") for s in row_strs):
            header_idx = i
            break
    
    if header_idx is None:
        wb.close()
        return {"branch": "", "period": "", "operators": [], "jobs": []}
    
    headers = [str(h or "").strip() for h in rows[header_idx]]
    
    # Build column map using .includes() style matching
    def ci(*candidates):
        for c in candidates:
            for idx, h in enumerate(headers):
                if c.lower() in h.lower():
                    return idx
        return -1
    
    id_col     = ci("Shipment ID", "Job Number")
    dept_col   = ci("Job Dept", "Department")
    status_col = ci("Job Status", "Invoice Status", "Job Stat")
    op_col     = ci("Job Operator", "Operator", "Job Ops", "Ops", "Sales")
    branch_col = ci("Branch", "Job Brn.", "Br.", "Br")
    consignor  = ci("Consignor")
    consignee  = ci("Consignee")
    origin_col = ci("Origin")
    dest_col   = ci("Destination", "Dest.")
    etd_col    = ci("Origin ETD", "ETD", "Job Open")
    eta_col    = ci("Destination ETA", "ETA")
    rev_col    = ci("Total Revenue", "Revenue", "Local Rev", "OS Amount")
    wip_col    = ci("Total WIP", "WIP")
    accr_col   = ci("Total Accrual", "Accrual")
    cost_col   = ci("Total Cost", "Cost", "Cost Local")
    profit_col = ci("Job Profit", "Profit")
    local_client_col = ci("Local Client", "Client", "Consignor")
    
    all_jobs = []
    operators_set = set()
    
    for row in rows[header_idx + 1:]:
        if not row or not row[id_col]:
            continue
        
        job_id = str(row[id_col] or "").strip()
        if not job_id:
            continue
        
        def _g(idx):
            return row[idx] if idx >= 0 and idx < len(row) else None
        
        dept = str(_g(dept_col) or "").strip()
        op_code = str(_g(op_col) or "").strip()
        if not op_code:
            op_code = "UNASSIGNED"
        op = OPERATOR_NAMES.get(op_code, op_code)
        open_date_raw = _g(etd_col) if is_export_dept(dept) else _g(eta_col)
        if not open_date_raw:
            open_date_raw = _g(etd_col) or _g(eta_col)
        
        if op:
            operators_set.add(op)
            
        branch_code = str(_g(branch_col) or branch).strip()
        job_branch = normalize_branch_name(branch_code)
        
        job = {
            "job_number":     job_id,
            "job_status":     str(_g(status_col) or "").strip(),
            "branch":         job_branch,

            "department":     dept,
            "open_date":      _parse_date_str(open_date_raw),
            "operator":       op,
            "sales_rep":      "",
            "local_charges":  "",
            "overseas_agent": "",
            "local_client":   str(_g(local_client_col) or "").strip(),
            "revenue":        _parse_number(_g(rev_col)),
            "wip":            _parse_number(_g(wip_col)),
            "cost":           _parse_number(_g(cost_col)),
            "accrual":        _parse_number(_g(accr_col)),
            "profit_loss":    _parse_number(_g(profit_col)),
            "margin_pct":     0.0,
            "job_age_days":   _compute_age(open_date_raw),
            "is_export":      is_export_dept(dept),
            "origin":         str(_g(origin_col) or "").strip(),
            "destination":    str(_g(dest_col) or "").strip(),
            "etd":            _parse_date_str(_g(etd_col)),
            "eta":            _parse_date_str(_g(eta_col)),
            "source_type":    source_type,
        }
        
        # If no Profit column in file, compute P/L from Revenue - abs(Cost)
        if profit_col < 0 and job["revenue"] != 0:
            job["profit_loss"] = round(job["revenue"] - abs(job["cost"]), 2)
        
        # Calculate margin
        if job["revenue"] != 0 and job["profit_loss"] != 0:
            job["margin_pct"] = round((job["profit_loss"] / job["revenue"]) * 100, 2)
        
        all_jobs.append(job)
        
    period = datetime.now().strftime("%B %Y")
    valid_dates = [j["open_date"] for j in all_jobs if j.get("open_date")]
    if valid_dates:
        try:
            # Try parsing DD/MM/YYYY
            dates = [datetime.strptime(d, "%d/%m/%Y") for d in valid_dates]
            max_date = max(dates)
            period = max_date.strftime("%B %Y")
        except ValueError:
            pass

    for job in all_jobs:
        job["flags"] = get_flags(job, period)
        job["primary_flag"] = priority_flag(job["flags"])
        job["ops_section"] = get_ops_section(job)
    
    wb.close()
    
    return {
        "branch": BRANCH_NAMES.get(branch, branch) if branch else "ALL",
        "period": period,
        "operators": sorted(operators_set),
        "jobs": all_jobs,
    }


def parse_job_billing(file_bytes: bytes, is_aged_file: bool = False) -> dict:
    """Parse 'Job Billing - Charges Not Yet Posted as REV or CST' or Aged Accruals report."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    
    header_idx = -1
    for i, row in enumerate(rows[:30]):
        row_strs = [str(c or "").strip().lower() for c in row]
        if "job number" in row_strs and "cost local" in row_strs:
            header_idx = i
            break
            
    if header_idx == -1:
        wb.close()
        return {"branch": "", "period": "", "operators": [], "jobs": []}
        
    headers = [str(h or "").strip().lower() for h in rows[header_idx]]
    
    id_col = headers.index("job number") if "job number" in headers else -1
    stat_col = headers.index("job stat") if "job stat" in headers else -1
    branch_col = headers.index("job brn.") if "job brn." in headers else -1
    dept_col = headers.index("job dept") if "job dept" in headers else -1
    op_col = headers.index("job ops") if "job ops" in headers else -1
    date_col = headers.index("job open") if "job open" in headers else -1
    cost_col = headers.index("cost local") if "cost local" in headers else -1
    client_col = headers.index("local client") if "local client" in headers else -1
    
    charge_col = headers.index("charge code") if "charge code" in headers else -1
    os_cur_col = headers.index("os cur") if "os cur" in headers else -1
    os_amt_col = headers.index("os amount") if "os amount" in headers else -1
    ex_rate_col = headers.index("ex rate") if "ex rate" in headers else -1
    creditor_col = headers.index("creditor") if "creditor" in headers else -1
    has_acr_col = headers.index("has acr") if "has acr" in headers else -1
    acr_rec_col = headers.index("acr recognised") if "acr recognised" in headers else -1
    
    job_accruals = {}
    operators_set = set()
    branch = ""
    
    for row in rows[header_idx + 1:]:
        if not row or not row[id_col]:
            continue
            
        job_id = str(row[id_col]).strip()
        if not job_id or job_id == "None":
            continue
            
        def _g(idx): return row[idx] if idx >= 0 and idx < len(row) else None
        
        cost = _parse_number(_g(cost_col))
        if job_id not in job_accruals:
            op_code = str(_g(op_col) or "").strip() or "UNASSIGNED"
            op = OPERATOR_NAMES.get(op_code, op_code)
            dept = str(_g(dept_col) or "").strip()
            branch_code = str(_g(branch_col) or "").strip()
            open_date_raw = _g(date_col)
            
            if op: operators_set.add(op)
            if branch_code and not branch: branch = branch_code
            
            job_accruals[job_id] = {
                "job_number": job_id,
                "job_status": str(_g(stat_col) or "").strip(),
                "branch": BRANCH_NAMES.get(branch_code, branch_code),
                "department": dept,
                "open_date": _parse_date_str(open_date_raw),
                "operator": op,
                "sales_rep": "",
                "local_charges": "",
                "overseas_agent": "",
                "local_client": str(_g(client_col) or "").strip(),
                "has_aged_accruals": is_aged_file,
                "revenue": 0.0,
                "wip": 0.0,
                "cost": 0.0,
                "accrual": 0.0,
                "profit_loss": 0.0,
                "margin_pct": 0.0,
                "job_age_days": 0,
                "is_export": is_export_dept(dept),
                "origin": "",
                "destination": "",
                "etd": _parse_date_str(open_date_raw),
                "eta": _parse_date_str(open_date_raw),
                "accrual_lines": [],
            }
            
        # Accumulate the "Cost Local" into accrual
        job_accruals[job_id]["accrual"] += cost
        
        # Build line item and compute age from ACR Recognised
        acr_rec_raw = _g(acr_rec_col)
        open_date_raw = _g(date_col)
        line_age = _compute_age(acr_rec_raw) if acr_rec_raw else _compute_age(open_date_raw)
        
        line_item = {
            "charge_code": str(_g(charge_col) or "").strip() if charge_col >= 0 else "",
            "os_cur": str(_g(os_cur_col) or "").strip() if os_cur_col >= 0 else "",
            "os_amount": _parse_number(_g(os_amt_col)) if os_amt_col >= 0 else 0.0,
            "ex_rate": _parse_number(_g(ex_rate_col)) if ex_rate_col >= 0 else 1.0,
            "cost_local": cost,
            "creditor": str(_g(creditor_col) or "").strip() if creditor_col >= 0 else "",
            "has_acr": str(_g(has_acr_col) or "").strip() if has_acr_col >= 0 else "",
            "acr_recognised": _parse_date_str(acr_rec_raw),
            "age_days": line_age,
        }
        job_accruals[job_id]["accrual_lines"].append(line_item)
        job_accruals[job_id]["job_age_days"] = max(job_accruals[job_id]["job_age_days"], line_age)

    all_jobs = []
    period = datetime.now().strftime("%B %Y")
    
    for job in job_accruals.values():
        job["flags"] = get_flags(job, period)
        job["primary_flag"] = priority_flag(job["flags"])
        job["ops_section"] = get_ops_section(job)
        all_jobs.append(job)
        
    wb.close()
    return {
        "branch": BRANCH_NAMES.get(branch, branch) if branch else "ALL",
        "period": period,
        "operators": sorted(operators_set),
        "jobs": all_jobs,
    }



def is_neg_movement_file(file_bytes: bytes, filename: str = "") -> bool:
    fn = filename.lower()
    if "negative" in fn or "movement" in fn:
        return True
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True)
        sheet_names = [s.lower() for s in wb.sheetnames]
        wb.close()
        if any("negative" in s or "excess profit" in s or "losses" in s for s in sheet_names):
            return True
    except Exception:
        pass
    return False



def parse_excel(file_bytes: bytes, filename: str = "") -> dict:
    """Auto-detect format and parse."""
    if is_neg_movement_file(file_bytes, filename):
        return {"branch": "", "period": "", "operators": [], "jobs": []}

    fn = filename.lower()
    if "wip_review" in fn or "wip review" in fn:
        return parse_wip_review(file_bytes)
    
    is_aged = "aged" in fn or ("accrual" in fn and ("3 month" in fn or "greater" in fn))
    if "billing" in fn or "charges" in fn or "not yet posted" in fn or "accrual" in fn or "aged" in fn:
        return parse_job_billing(file_bytes, is_aged_file=is_aged)
    
    # Detect source_type from filename for pending invoicing files
    source_type = ""
    if "cross" in fn and "trade" in fn:
        source_type = "cross_trade"
    elif "export" in fn:
        source_type = "exports"
    elif "import" in fn and " b " in fn or fn.endswith(" b.xlsx") or "import b" in fn:
        source_type = "imports_b"
    elif "import" in fn and (" s " in fn or fn.endswith(" s.xlsx") or "import s" in fn):
        source_type = "imports_s"
    
    # Try to detect by checking sheet names
    wb = load_workbook(BytesIO(file_bytes), read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    
    # If it has a "Shipment Profile" sheet, it's CargoWise
    if any("shipment" in s.lower() for s in sheet_names):
        return parse_cargowise_export(file_bytes, source_type=source_type)
    
    # Check if first sheet has 'job number' and 'cost local' (Job Billing / Aged Accruals format)
    try:
        wb_check = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        ws_check = wb_check.active
        for row in list(ws_check.iter_rows(values_only=True))[:30]:
            row_s = [str(c or "").strip().lower() for c in row]
            if "job number" in row_s and "cost local" in row_s:
                wb_check.close()
                return parse_job_billing(file_bytes)
        wb_check.close()
    except Exception:
        pass
    
    # If it has named operator sheets (2-3 char codes), treat as WIP Review
    if any(len(s) <= 4 and s.isalnum() for s in sheet_names):
        return parse_wip_review(file_bytes)
    
    # Default to CargoWise
    return parse_cargowise_export(file_bytes, source_type=source_type)
