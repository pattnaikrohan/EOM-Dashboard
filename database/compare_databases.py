"""
Compare tables across DEV and PROD databases.
Output an Excel file showing what's in DEV but missing from PROD.
"""
import snowflake.connector
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.backends import default_backend
import json

SF_ACCOUNT   = "SGLYREN-GG43054"
SF_USER      = "TEST_AI_AUTO"
SF_WAREHOUSE = "PROD_COMPUTE_WH"
SF_ROLE      = "PROD_ENGINEER"

PRIVATE_KEY_PEM = """-----BEGIN PRIVATE KEY-----
MIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQDFkLT/bHavGuUi
y/7e+nae+Pyitn5Y7ZFpct73cZnGK18+YD7MnexY4vYFNgb7Io9GEEnPUcUk0N0a
HzUwUcHMey/tMRxDzGB2mj3Vqgr1HkwEpwckOjnJiqoo29IqNcFSzGhCTnuBQAS9
u1y49pRiXXBQNzHN2m1CP/CTI6MlL0l/QZ4osQhBb41avjcl3c0iQzn/OxGvUaUf
cF30cWQWiv3fQ8lXf29/uA8UIWF1Wp/O2bEJgqF6BX1tGcml2QrkZBgIdURJpYgg
CccywefO2MkfIf7vzttcWSA2dByn8ImmcBo1t2dIfaERa+ZNGNfEcGFEpBGGo0F4
uGHHZGHPAgMBAAECggEAHUT1tmxt85NFzy9EcJR6w81E4HZY/E519Ja9dxCwJNkt
dc2cgC6wDUXSslhGwH4Q6+MN4avOS/Na+REWjYoPRClxTZqxAqUMxYV5vmXq1FXL
aAUMMc8j2zqylaBmNTVcPCq8Z4ETxcT47thglvvbupb372ponoQgcRLipBSWNF6B
9PegOKGhGeJ1APb693Lr9hkFvM8mq/usoQMe1sLvc+AbSV7O0Sd7fs0bn9TkWaxC
VA1QyiVtCBspfwtEuXpByrshgxbkfszC0R3mGn+Ojcpihbk+/s7mfdH+Ep72QPN6
IFg0SOHUA4HwR6RkDsPLdX4oSccaRropYP3mTYD1QQKBgQD+JGx06OzX6AOsuLas
OhcXVFw2rZ5kggFnzTksRSXBwJBINDclK+Ah4Lq6jbqap84pzOPJe2Xyx1nQ6b6q
VqYO/tcVN3BgpHxxNOc1YvSIEeP2gy97AuVzD+5bFvgF7x5Bml4mgRQIb0uYfEby
b0UKlGxodyelt7MWF8tNRBTkbQKBgQDHAmkrEgQOJ/g/CjHWDcqywQt5JVQGP3ZC
5Im+C37v/WZge77fg2oYpv0yJ1QIuLiCp8MDRYdLahru0/JUsL8Yu6gb9tXc0MpE
vxoVJBiCDJfdOSHZs+AOmUDnLBQBoZLHxUbGC6ltwRALSHQGRmH7EPdN4UJg5awJ
MehRT9zhqwKBgQDiq/AEMkAUrj6gzjs3z7QKvdZlemM8t+uy/osQ3je34R2PGOta
fxCwhrVlcMXP7P1nsPQ2H5alfIKyX9kMKq5/z3Jc3Q6hU/QeMJZLuo/p0TMnCojN
yZ6HCt2IZysed9DfqGRzKUuJ3mJphebtkqrcrdcnMaeuGfNkMCLHLoMH3QKBgQC5
AS+9n4DvnA62pAaSZL3kEXxWAfKr4EFTjFvUtaEq/5o15bQa23M9Obg18MO5W+gD
ZmvvVaqh3CDvl083lhwApStx27UTE3KGGFXqA2VZONXRDbS/Su3nBGeGwL5Uid0H
JAlYQS0f4BPHOXLLBpE9spcE6n0n0TtuTNwZAzJnAwKBgGnb/uagqCameqnDoVrw
RKdz/6DLX6Ul3ql5FU/KH9zKH5yKa8arOONJKqggrF920vaLZuArAMNL7e1KSbMv
FY+Se1JzkprZUnYRrZDyJy1xUzfpPmfHdXJcr+7qNXWUgvVEP5pGR38YSjA6MDac
8QOHXsEQQry1/656UObS4Nvl
-----END PRIVATE KEY-----"""

def get_connection():
    p_key = serialization.load_pem_private_key(
        PRIVATE_KEY_PEM.encode(), password=None, backend=default_backend())
    pkb = p_key.private_bytes(
        encoding=serialization.Encoding.DER,
        format=serialization.PrivateFormat.PKCS8,
        encryption_algorithm=serialization.NoEncryption())
    return snowflake.connector.connect(
        account=SF_ACCOUNT, user=SF_USER, private_key=pkb,
        warehouse=SF_WAREHOUSE, role=SF_ROLE)

def q(cur, sql):
    try:
        cur.execute(sql)
        rows = cur.fetchall()
        cols = [d[0] for d in cur.description] if cur.description else []
        return rows, cols
    except Exception as e:
        return None, str(e)

def main():
    conn = get_connection()
    cur = conn.cursor()
    print("CONNECTED\n")

    # ── Collect ALL tables/views from DEV (all schemas) ───────────────────
    print("Fetching DEV tables...")
    dev_tables = {}
    for schema in ['CORE', 'RAW', 'PUBLIC']:
        rows, _ = q(cur, f"""
            SELECT TABLE_NAME, TABLE_TYPE, ROW_COUNT, BYTES
            FROM DEV.INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = '{schema}'
            ORDER BY TABLE_NAME
        """)
        if rows:
            for r in rows:
                key = f"{schema}.{r[0]}"
                dev_tables[key] = {
                    'schema': schema,
                    'table_name': r[0],
                    'table_type': r[1],
                    'row_count': r[2],
                    'bytes': r[3]
                }
    print(f"  DEV: {len(dev_tables)} objects found")

    # ── Collect ALL tables/views from PROD (all schemas) ──────────────────
    print("Fetching PROD tables...")
    prod_tables = {}
    for schema in ['CORE', 'RAW', 'EDW', 'MART', 'PUBLIC']:
        rows, _ = q(cur, f"""
            SELECT TABLE_NAME, TABLE_TYPE, ROW_COUNT, BYTES
            FROM PROD.INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = '{schema}'
            ORDER BY TABLE_NAME
        """)
        if rows:
            for r in rows:
                key = f"{schema}.{r[0]}"
                prod_tables[key] = {
                    'schema': schema,
                    'table_name': r[0],
                    'table_type': r[1],
                    'row_count': r[2],
                    'bytes': r[3]
                }
    print(f"  PROD: {len(prod_tables)} objects found")

    # ── Also check UAT ────────────────────────────────────────────────────
    print("Fetching UAT tables...")
    uat_tables = {}
    rows, _ = q(cur, "SHOW SCHEMAS IN DATABASE UAT")
    uat_schemas = [r[1] for r in rows] if rows else []
    for schema in uat_schemas:
        if schema == 'INFORMATION_SCHEMA':
            continue
        rows, _ = q(cur, f"""
            SELECT TABLE_NAME, TABLE_TYPE, ROW_COUNT, BYTES
            FROM UAT.INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = '{schema}'
            ORDER BY TABLE_NAME
        """)
        if rows:
            for r in rows:
                key = f"{schema}.{r[0]}"
                uat_tables[key] = {
                    'schema': schema,
                    'table_name': r[0],
                    'table_type': r[1],
                    'row_count': r[2],
                    'bytes': r[3]
                }
    print(f"  UAT: {len(uat_tables)} objects found")

    # ── Build comparison data ─────────────────────────────────────────────
    # Get all unique table names from DEV
    dev_table_names = set(t['table_name'] for t in dev_tables.values())
    prod_table_names = set(t['table_name'] for t in prod_tables.values())
    uat_table_names = set(t['table_name'] for t in uat_tables.values())

    # Save as JSON for the Excel generator
    comparison = []
    for key, info in sorted(dev_tables.items()):
        tname = info['table_name']
        # Check if this table exists in PROD (any schema)
        in_prod = tname in prod_table_names
        in_uat = tname in uat_table_names

        # Find which PROD schemas it's in
        prod_schemas = [t['schema'] for t in prod_tables.values() if t['table_name'] == tname]
        uat_schemas_found = [t['schema'] for t in uat_tables.values() if t['table_name'] == tname]
        
        prod_rows = sum([t['row_count'] for t in prod_tables.values() if t['table_name'] == tname and t['row_count'] is not None]) if in_prod else None
        uat_rows = sum([t['row_count'] for t in uat_tables.values() if t['table_name'] == tname and t['row_count'] is not None]) if in_uat else None

        comparison.append({
            'dev_schema': info['schema'],
            'table_name': tname,
            'table_type': info['table_type'],
            'dev_rows': info['row_count'],
            'dev_bytes': info['bytes'],
            'in_prod': in_prod,
            'prod_schemas': ', '.join(prod_schemas) if prod_schemas else '-',
            'prod_rows': prod_rows,
            'in_uat': in_uat,
            'uat_schemas': ', '.join(uat_schemas_found) if uat_schemas_found else '-',
            'uat_rows': uat_rows,
            'status': 'MISSING FROM PROD' if not in_prod else 'EXISTS IN PROD'
        })

    # Save JSON
    with open(r'd:\EOM DASHBOARDS PROTO\database\table_comparison.json', 'w') as f:
        json.dump(comparison, f, indent=2, default=str)

    # ── Print summary ─────────────────────────────────────────────────────
    missing = [c for c in comparison if not c['in_prod']]
    present = [c for c in comparison if c['in_prod']]

    print(f"\n{'='*80}")
    print(f"SUMMARY")
    print(f"{'='*80}")
    print(f"  DEV tables/views total:     {len(comparison)}")
    print(f"  Present in PROD:            {len(present)}")
    print(f"  MISSING from PROD:          {len(missing)}")
    print(f"\nMISSING FROM PROD:")
    print(f"{'Table Name':50s} {'DEV Schema':12s} {'Type':12s} {'DEV Rows':>12s}")
    print(f"{'-'*50} {'-'*12} {'-'*12} {'-'*12}")
    for c in sorted(missing, key=lambda x: x['table_name']):
        rows_str = f"{c['dev_rows']:,}" if c['dev_rows'] is not None else 'N/A'
        print(f"{c['table_name']:50s} {c['dev_schema']:12s} {c['table_type']:12s} {rows_str:>12s}")

    cur.close()
    conn.close()
    print("\nJSON saved. Generating Excel...")

    # ── Generate Excel ────────────────────────────────────────────────────
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Sheet 1: Missing from PROD ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Missing from PROD"

    # Header style
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    headers1 = ['#', 'Table Name', 'DEV Schema', 'Table Type', 'DEV Row Count',
                 'In UAT?', 'UAT Schema', 'Notes']
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data
    row_fill_alt = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    data_font = Font(name='Calibri', size=10)

    for i, c in enumerate(sorted(missing, key=lambda x: x['table_name']), 1):
        row = i + 1
        vals = [
            i,
            c['table_name'],
            c['dev_schema'],
            c['table_type'],
            c['dev_rows'],
            'Yes' if c['in_uat'] else 'No',
            c['uat_schemas'],
            'Required for EOM Dashboard' if c['table_name'] in [
                'JOBCHARGE','JOBHEADER','ACCCHARGECODE','JOBCHARGEREVRECOGNITION',
                'GLBSTAFF','ORGHEADER','ORGADDRESS','JOBSHIPMENT','JOBDECLARATION',
                'GLBBRANCH'] else ''
        ]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=col, value=v)
            cell.font = data_font
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = row_fill_alt
            if col == 5 and v is not None:
                cell.number_format = '#,##0'

    # Column widths
    widths1 = [5, 50, 12, 12, 15, 10, 12, 30]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Full Comparison ──────────────────────────────────────────
    ws2 = wb.create_sheet("Full Comparison")

    green_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    red_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
    alt_fill = PatternFill(start_color='F2F3F4', end_color='F2F3F4', fill_type='solid')

    headers2 = ['#', 'Table Name', 'DEV Schema', 'Table Type', 'DEV Row Count',
                 'In PROD?', 'PROD Schema(s)', 'PROD Row Count', 'In UAT?', 'UAT Schema(s)', 'UAT Row Count', 'Status']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        cell.alignment = header_align
        cell.border = thin_border

    for i, c in enumerate(sorted(comparison, key=lambda x: (x['status'], x['table_name'])), 1):
        row = i + 1
        vals = [
            i,
            c['table_name'],
            c['dev_schema'],
            c['table_type'],
            c['dev_rows'],
            'Yes' if c['in_prod'] else 'NO',
            c['prod_schemas'],
            c['prod_rows'],
            'Yes' if c['in_uat'] else 'No',
            c['uat_schemas'],
            c['uat_rows'],
            c['status']
        ]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col, value=v)
            cell.font = data_font
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = alt_fill
            if col in (5, 8, 11) and v is not None:
                cell.number_format = '#,##0'
            # Highlight status
            if col == 12:
                if 'MISSING' in str(v):
                    cell.font = Font(name='Calibri', size=10, bold=True, color='C0392B')
                else:
                    cell.font = Font(name='Calibri', size=10, color='27AE60')
            if col == 6 and v == 'NO':
                cell.font = Font(name='Calibri', size=10, bold=True, color='C0392B')

    widths2 = [5, 50, 12, 12, 15, 10, 15, 15, 10, 15, 15, 22]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: EOM Dashboard Required Tables ────────────────────────────
    ws3 = wb.create_sheet("EOM Required Tables")

    eom_tables = ['JOBCHARGE','JOBHEADER','ACCCHARGECODE','JOBCHARGEREVRECOGNITION',
                  'GLBSTAFF','ORGHEADER','ORGADDRESS','JOBSHIPMENT','JOBDECLARATION','GLBBRANCH']

    headers3 = ['#', 'Table Name', 'Purpose', 'DEV.CORE Rows', 'DEV.CORE Unique PKs',
                 'In PROD?', 'PROD Schema', 'Action Needed']
    purposes = {
        'JOBCHARGE': 'Charge lines (cost/sell per line)',
        'JOBHEADER': 'Job master (status, direction, dates)',
        'ACCCHARGECODE': 'Charge code reference + GL accounts',
        'JOBCHARGEREVRECOGNITION': 'WIP release/recognition events',
        'GLBSTAFF': 'Staff names (Operator, Sales Rep)',
        'ORGHEADER': 'Organisation names (Agent, Client)',
        'ORGADDRESS': 'Address bridge (OA_PK -> OH_PK)',
        'JOBSHIPMENT': 'Forwarding routing (origin/dest/ETD/ETA)',
        'JOBDECLARATION': 'Customs routing (origin/dest/dates)',
        'GLBBRANCH': 'Branch code/name lookup',
    }
    unique_pks = {
        'JOBCHARGE': 375035, 'JOBHEADER': 43629, 'ACCCHARGECODE': 277,
        'JOBCHARGEREVRECOGNITION': 35825, 'GLBSTAFF': 187, 'ORGHEADER': 15917,
        'ORGADDRESS': 10061, 'JOBSHIPMENT': 39240, 'JOBDECLARATION': 37207,
        'GLBBRANCH': 11,
    }

    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
        cell.alignment = header_align
        cell.border = thin_border

    for i, tname in enumerate(eom_tables, 1):
        row = i + 1
        in_prod = tname in prod_table_names
        prod_s = [t['schema'] for t in prod_tables.values() if t['table_name'] == tname]
        dev_info = next((t for t in dev_tables.values() if t['table_name'] == tname and t['schema'] == 'CORE'), None)

        vals = [
            i,
            tname,
            purposes.get(tname, ''),
            dev_info['row_count'] if dev_info else 'N/A',
            unique_pks.get(tname, 'N/A'),
            'Yes' if in_prod else 'NO',
            ', '.join(prod_s) if prod_s else '-',
            'None' if in_prod else 'NEEDS REPLICATION TO PROD'
        ]
        for col, v in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=col, value=v)
            cell.font = data_font
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = alt_fill
            if col in (4, 5) and isinstance(v, (int, float)):
                cell.number_format = '#,##0'
            if col == 8 and 'NEEDS' in str(v):
                cell.font = Font(name='Calibri', size=10, bold=True, color='C0392B')
            if col == 6 and v == 'NO':
                cell.font = Font(name='Calibri', size=10, bold=True, color='C0392B')

    widths3 = [5, 30, 40, 15, 18, 10, 15, 30]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Auto-filter on all sheets
    for ws in [ws1, ws2, ws3]:
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'

    # Save
    output_path = r'd:\EOM DASHBOARDS PROTO\database\DEV_vs_PROD_Table_Comparison_v2.xlsx'
    wb.save(output_path)
    print(f"\nExcel saved to: {output_path}")
    print("DONE")

if __name__ == "__main__":
    main()
